#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
import sys
import re
from datetime import datetime, date
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


DELIM = "|"  # 複数値の区切り文字


def to_str(v: Any) -> str:
    """Excelセル値を文字列に寄せる（Noneは空文字）。"""
    if v is None:
        return ""
    # openpyxlは日付型をdatetime/dateで返すことがある
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    # 数値IDなどが入っても文字列化
    return str(v).strip()


def split_list(s: str) -> List[str]:
    """'a| b |c' -> ['a','b','c'] / 空なら[]"""
    s = (s or "").strip()
    if not s:
        return []
    parts = [p.strip() for p in s.split(DELIM)]
    return [p for p in parts if p]  # 空要素除去


def is_probably_url(s: str) -> bool:
    return bool(re.match(r"^https?://", s.strip(), re.IGNORECASE))


def normalize_header(h: str) -> str:
    """ヘッダ名を最低限正規化（前後空白除去・小文字化）"""
    return (h or "").strip()


def load_sheet(path: Path, sheet_name: str = "Terms"):
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. sheets={wb.sheetnames}")
    return wb[sheet_name]


def build_header_map(ws) -> Dict[str, int]:
    """1行目の列名→列index(1-based)"""
    header_map: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        key = normalize_header(to_str(val))
        if key:
            header_map[key] = col
    return header_map


def get_cell(ws, row: int, header_map: Dict[str, int], key: str) -> Any:
    col = header_map.get(key)
    if not col:
        return None
    return ws.cell(row=row, column=col).value


def row_is_empty(ws, row: int, header_map: Dict[str, int]) -> bool:
    # id が空なら「空行扱い」とする（途中スキップのため）
    v = to_str(get_cell(ws, row, header_map, "id"))
    return v == ""


def validate_required(term_id: str, term: str) -> Tuple[bool, str]:
    if not term_id:
        return False, "missing id"
    if not term:
        return False, "missing term"
    return True, ""


def convert_excel_to_terms(
    xlsx_path: Path,
    out_json_path: Path,
    sheet_name: str = "Terms",
    stop_on_blank_id: bool = False,
) -> Dict[str, Any]:
    ws = load_sheet(xlsx_path, sheet_name=sheet_name)
    header_map = build_header_map(ws)

    # 必須列チェック
    for required in ["id", "term"]:
        if required not in header_map:
            raise ValueError(f"Required column '{required}' not found in header row.")

    terms: List[Dict[str, Any]] = []
    warnings: List[str] = []
    seen_ids = set()

    max_row = ws.max_row
    for r in range(2, max_row + 1):
        term_id = to_str(get_cell(ws, r, header_map, "id"))
        if not term_id:
            if stop_on_blank_id:
                break
            else:
                continue

        term = to_str(get_cell(ws, r, header_map, "term"))
        ok, reason = validate_required(term_id, term)
        if not ok:
            warnings.append(f"Row {r}: skip ({reason})")
            continue

        if term_id in seen_ids:
            warnings.append(f"Row {r}: duplicate id '{term_id}' (skip)")
            continue
        seen_ids.add(term_id)

        reading = to_str(get_cell(ws, r, header_map, "reading"))
        en = to_str(get_cell(ws, r, header_map, "en"))

        category = split_list(to_str(get_cell(ws, r, header_map, "category")))
        tags = split_list(to_str(get_cell(ws, r, header_map, "tags")))

        summary = to_str(get_cell(ws, r, header_map, "summary"))
        body = to_str(get_cell(ws, r, header_map, "body"))

        related_ids_raw = to_str(get_cell(ws, r, header_map, "related_ids"))
        related = split_list(related_ids_raw)

        source_raw = to_str(get_cell(ws, r, header_map, "source"))
        source = split_list(source_raw)

        owner = to_str(get_cell(ws, r, header_map, "owner"))
        status = to_str(get_cell(ws, r, header_map, "status")) or "draft"
        updated = to_str(get_cell(ws, r, header_map, "updated"))
        created = to_str(get_cell(ws, r, header_map, "created"))

        # status正規化（任意だが安全策）
        if status not in {"draft", "verified", "deprecated"}:
            warnings.append(f"Row {r} id={term_id}: unknown status '{status}' -> set 'draft'")
            status = "draft"

        # 日付フォーマット軽いチェック（空は許容）
        for k, v in [("updated", updated), ("created", created)]:
            if v and not re.match(r"^\d{4}-\d{2}-\d{2}$", v):
                warnings.append(f"Row {r} id={term_id}: {k} not YYYY-MM-DD ('{v}')")

        # relatedの存在チェック（後でリンク切れになるため警告）
        # ここでは後段でまとめてチェックするので、まず保持

        item = {
            "id": term_id,
            "term": term,
            "reading": reading,
            "en": en,
            "category": category,
            "tags": tags,
            "summary": summary,
            "body": body,
            "related": related,
            "source": source,
            "owner": owner,
            "status": status,
            "updated": updated,
            "created": created,
        }
        terms.append(item)

    # relatedのリンク整合性チェック
    id_set = {t["id"] for t in terms}
    for t in terms:
        bad = [rid for rid in t.get("related", []) if rid and rid not in id_set]
        if bad:
            warnings.append(f"id={t['id']}: related id not found {bad}")

    out_json_path.parent.mkdir(parents=True, exist_ok=True)
    with out_json_path.open("w", encoding="utf-8") as f:
        json.dump(terms, f, ensure_ascii=False, indent=2)

    report = {
        "input": str(xlsx_path),
        "sheet": sheet_name,
        "output": str(out_json_path),
        "count": len(terms),
        "warnings": warnings,
    }
    return report


def main():
    import argparse

    p = argparse.ArgumentParser(description="Convert glossary Excel to terms.json")
    p.add_argument("--xlsx", default="terms.xlsx", help="Input Excel path (default: terms.xlsx)")
    p.add_argument("--out", default="data/terms.json", help="Output JSON path (default: data/terms.json)")
    p.add_argument("--sheet", default="Terms", help="Sheet name (default: Terms)")
    p.add_argument(
        "--stop-on-blank-id",
        action="store_true",
        help="Stop processing when a blank id row is found (instead of skipping).",
    )
    p.add_argument(
        "--report",
        default="data/convert_report.json",
        help="Report JSON path (default: data/convert_report.json)",
    )
    args = p.parse_args()

    report = convert_excel_to_terms(
        xlsx_path=Path(args.xlsx),
        out_json_path=Path(args.out),
        sheet_name=args.sheet,
        stop_on_blank_id=args.stop_on_blank_id,
    )

    # レポート出力（警告確認用）
    report_path = Path(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Converted: {report['count']} terms -> {report['output']}")
    if report["warnings"]:
        print(f"Warnings: {len(report['warnings'])} (see {report_path})")
    else:
        print("Warnings: 0")


if __name__ == "__main__":
    main()
